from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

OUTPUT = Path('/home/ubuntu/SESA-DASH/database/sesa-mysql-data-import-template.xlsx')
NAVY = '123B5D'
BLUE = '1D6A96'
PALE = 'EAF2F7'
GRAY = '5B6770'
BORDER = Border(bottom=Side(style='thin', color='B7C9D6'))

SHEETS = [
    {
        'name': '01_Departments',
        'title': 'DEPARTMANLAR / İŞ MERKEZLERİ',
        'table': 'departments',
        'columns': [
            ('id', 'İş Merkezi ID / Kodu', 'Zorunlu; SESA-DASH ilişki anahtarı. Canias iş merkezi kodunu kullanın.'),
            ('name', 'İş Merkezi / Departman Adı', 'Zorunlu; örn. Kesim, Baskı, Laminasyon.'),
            ('canias_work_center_code', 'İş Merkezi Kodu (Canias)', 'Canias’taki benzersiz iş merkezi kodu.'),
            ('canias_department_code', 'Departman Kodu (Canias)', 'Canias departman kodu.'),
            ('is_active', 'Aktif mi?', '1 aktif, 0 pasif.'),
        ],
        'validations': {'Aktif mi?': '"1,0"'},
    },
    {
        'name': '02_Users',
        'title': 'KULLANICILAR / PERSONEL',
        'table': 'users',
        'columns': [
            ('personnel_no', 'Personel Sicil No (Canias)', 'Canias’taki benzersiz personel sicil numarası.'),
            ('full_name', 'Ad Soyad', 'Zorunlu.'),
            ('email', 'Kurumsal E-posta', 'Zorunlu; benzersiz olmalı.'),
            ('role', 'Rol', 'admin, operator, maintenance veya sorumlu.'),
            ('password_hash', 'Şifre Hash', 'Elle yazılmaz. Kullanıcı ilk kayıt/giriş akışından oluşturulmalıdır.'),
            ('is_active', 'Aktif mi?', '1 aktif, 0 pasif.'),
        ],
        'validations': {'Rol': '"admin,operator,maintenance,sorumlu"', 'Aktif mi?': '"1,0"'},
    },
    {
        'name': '03_UserDepartments',
        'title': 'PERSONEL – İŞ MERKEZİ YETKİLERİ',
        'table': 'user_departments',
        'columns': [
            ('user_id', 'Personel ID', '02_Users sayfasındaki kullanıcı ID’si veya import sırasında eşleştirilecek anahtar.'),
            ('department_id', 'İş Merkezi ID / Kodu', '01_Departments sayfasındaki id değeri.'),
        ],
        'validations': {},
    },
    {
        'name': '04_Machines',
        'title': 'MAKİNELER',
        'table': 'machines',
        'columns': [
            ('code', 'Makine Kodu (Canias)', 'Zorunlu ve benzersiz; örn. KESIM-01.'),
            ('name', 'Makine Adı / Modeli', 'Zorunlu.'),
            ('department_id', 'İş Merkezi ID / Kodu', '01_Departments.id ile eşleşmeli.'),
            ('qr_code_value', 'QR Kodu / Değeri', 'QR içine yazılacak makine kodu veya URL; benzersiz olmalı.'),
            ('canias_asset_no', 'Canias Varlık / Demirbaş No', 'Varsa Canias varlık numarası.'),
            ('added_by_user_id', 'Ekleyen Personel ID', '02_Users.id; bilinmiyorsa boş bırakılabilir.'),
            ('is_active', 'Aktif mi?', '1 aktif, 0 pasif.'),
        ],
        'validations': {'Aktif mi?': '"1,0"'},
    },
    {
        'name': '05_Issues',
        'title': 'ARIZA / BAKIM KAYITLARI',
        'table': 'issues',
        'columns': [
            ('machine_id', 'Makine ID', '04_Machines import sonrası oluşan teknik ID; kodla eşleştirilir.'),
            ('machine_code', 'Makine Kodu', '04_Machines.code ile aynı olmalı.'),
            ('machine_name', 'Makine Adı', 'Olay anındaki makine adı.'),
            ('department_id', 'İş Merkezi ID / Kodu', 'Makinenin bağlı olduğu iş merkezi.'),
            ('reporter_user_id', 'Bildiren Personel ID', '02_Users.id; bilinmiyorsa import sırasında sicil no ile eşleştirilir.'),
            ('reporter_personnel_no', 'Bildiren Personel Sicil No', 'Canias sicil numarası.'),
            ('reporter_email', 'Bildiren E-posta', 'Kayıt anındaki e-posta.'),
            ('description', 'Arıza / Bakım Açıklaması', 'Zorunlu.'),
            ('photo_url', 'Fotoğraf URL', 'Varsa fotoğraf bağlantısı.'),
            ('status', 'Durum', 'Açık, İşlemde veya Çözüldü.'),
            ('resolved_at', 'Çözülme Tarihi', 'Çözüldüyse UTC tarih-saat; değilse boş.'),
            ('resolved_by_user_id', 'Çözen Personel ID', 'Çözen kullanıcı; çözülmediyse boş.'),
            ('created_at', 'Bildirim Tarihi', 'UTC tarih-saat; örn. 2026-08-18 09:30:00.'),
        ],
        'validations': {'Durum': '"Açık,İşlemde,Çözüldü"'},
    },
    {
        'name': '06_IssueHistory',
        'title': 'ARIZA DURUM GEÇMİŞİ',
        'table': 'issue_status_history',
        'columns': [
            ('issue_id', 'Arıza ID', '05_Issues import sonrası oluşan teknik ID.'),
            ('status', 'Durum', 'Açık, İşlemde veya Çözüldü.'),
            ('changed_by_user_id', 'Değiştiren Personel ID', '02_Users.id.'),
            ('changed_at', 'Değişiklik Tarihi', 'UTC tarih-saat.'),
        ],
        'validations': {'Durum': '"Açık,İşlemde,Çözüldü"'},
    },
    {
        'name': '07_AppSettings',
        'title': 'UYGULAMA AYARLARI',
        'table': 'app_settings',
        'columns': [
            ('setting_key', 'Ayar Anahtarı', 'Örn. uygulama adı veya bildirim ayarı.'),
            ('setting_value', 'Ayar Değeri', 'Ayar değeri.'),
            ('is_secret', 'Gizli mi?', 'Şifre, JWT veya bağlantı bilgisi yazmayın. 1 gizli, 0 normal.'),
        ],
        'validations': {'Gizli mi?': '"1,0"'},
    },
]

wb = Workbook()
wb.remove(wb.active)

start = wb.create_sheet('00_BAŞLANGIÇ')
start.sheet_view.showGridLines = False
start.merge_cells('A1:F1')
start['A1'] = 'SESA-DASH | MYSQL VERİ HAZIRLAMA VE AKTARIM DOSYASI'
start['A1'].font = Font(bold=True, color='FFFFFF', size=16)
start['A1'].fill = PatternFill('solid', fgColor=NAVY)
start['A1'].alignment = Alignment(horizontal='center')
start['A3'] = 'Bu dosya ne için kullanılır?'
start['A3'].font = Font(bold=True, color='FFFFFF', size=12)
start['A3'].fill = PatternFill('solid', fgColor=BLUE)
start.merge_cells('A3:F3')
start['A4'] = 'Şirketin gerçek CaniasERP personel, iş merkezi, makine ve arıza verilerini yazmak için kullanılır. Satır 5’ten itibaren veri girin; açıklama satırlarını silmeyin.'
start.merge_cells('A4:F4')
start['A4'].alignment = Alignment(wrap_text=True, vertical='top')
start.row_dimensions[4].height = 36
start.append([])
start.append(['Sıra', 'Excel Sayfası', 'MySQL Tablosu', 'Doldurulacak veri', 'Foreign key / ilişki', 'Zorunluluk'])
for cell in start[6]:
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = PatternFill('solid', fgColor=BLUE)
    cell.alignment = Alignment(wrap_text=True)
for index, item in enumerate(SHEETS, 1):
    relation = {
        '01_Departments': 'Ana referans tablosu',
        '02_Users': 'Ana personel tablosu',
        '03_UserDepartments': 'Users + Departments',
        '04_Machines': 'Departments + Users',
        '05_Issues': 'Machines + Users + Departments',
        '06_IssueHistory': 'Issues + Users',
        '07_AppSettings': 'Bağımsız ayar tablosu',
    }[item['name']]
    required = 'Önce doldur' if index <= 2 else 'İlişkili veriden sonra'
    start.append([index, item['name'], item['table'], item['title'], relation, required])
start.append([])
start.append(['DOLDURMA KURALI', 'Önce 01_Departments, sonra 02_Users, 03_UserDepartments, 04_Machines, 05_Issues ve 06_IssueHistory sırasını kullanın.', '', '', '', ''])
start.append(['GÜVENLİK', 'password_hash, JWT_SECRET, DATABASE_URL ve gerçek şifreleri Excel’e yazmayın. Kullanıcı parolaları uygulamanın kayıt/giriş akışıyla oluşturulmalıdır.', '', '', '', ''])
start.append(['TARİH', 'Tüm tarih-saat değerlerini UTC biçiminde YYYY-MM-DD HH:MM:SS yazın.', '', '', '', ''])
for col, width in {'A': 12, 'B': 24, 'C': 26, 'D': 42, 'E': 34, 'F': 24}.items():
    start.column_dimensions[col].width = width
start.freeze_panes = 'A7'

for item in SHEETS:
    ws = wb.create_sheet(item['name'])
    ws.sheet_view.showGridLines = False
    headers = [c[0] for c in item['columns']]
    technical = [c[1] for c in item['columns']]
    descriptions = [c[2] for c in item['columns']]
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws.cell(1, 1, item['title'])
    ws.cell(1, 1).font = Font(bold=True, color='FFFFFF', size=14)
    ws.cell(1, 1).fill = PatternFill('solid', fgColor=NAVY)
    ws.cell(1, 1).alignment = Alignment(horizontal='center')
    ws.append(headers)
    ws.append(technical)
    ws.append(descriptions)
    for row_num in (2, 3):
        for cell in ws[row_num]:
            cell.font = Font(bold=True, color='FFFFFF' if row_num == 2 else GRAY)
            cell.fill = PatternFill('solid', fgColor=BLUE if row_num == 2 else PALE)
            cell.alignment = Alignment(wrap_text=True, vertical='center')
            cell.border = BORDER
    for cell in ws[4]:
        cell.font = Font(italic=True, color=GRAY)
        cell.alignment = Alignment(wrap_text=True, vertical='top')
    ws.row_dimensions[4].height = 42
    ws.freeze_panes = 'A5'
    ws.auto_filter.ref = f'A2:{get_column_letter(len(headers))}104'
    for idx, header in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(idx)].width = max(20, min(40, len(header) + 10))
    for row in ws.iter_rows(min_row=5, max_row=104, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.number_format = '@'
    for label, formula in item['validations'].items():
        if label in headers:
            col = get_column_letter(headers.index(label) + 1)
            validation = DataValidation(type='list', formula1=formula, allow_blank=True)
            validation.error = 'Listeden geçerli bir değer seçin.'
            validation.errorTitle = 'Geçersiz değer'
            ws.add_data_validation(validation)
            validation.add(f'{col}5:{col}104')

wb.properties.title = 'SESA-DASH MySQL Veri Hazırlama ve Aktarım Şablonu'
wb.properties.creator = 'SESA-DASH'
OUTPUT.parent.mkdir(parents=True, exist_ok=True)
wb.save(OUTPUT)

check = load_workbook(OUTPUT, read_only=True)
expected = ['00_BAŞLANGIÇ'] + [item['name'] for item in SHEETS]
assert check.sheetnames == expected
print(OUTPUT)
