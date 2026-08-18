from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

OUTPUT = Path('/home/ubuntu/SESA-DASH/database/canias-data-entry-template.xlsx')
HEADER_FILL = PatternFill('solid', fgColor='123B5D')
SUBHEADER_FILL = PatternFill('solid', fgColor='1D6A96')
THIN_BLUE = Side(style='thin', color='B7C9D6')

sheets = [
    ('1_Departmanlar', 'DEPARTMAN / İŞ MERKEZİ VERİ GİRİŞİ', [
        ('İş Merkezi Kodu (Canias)', 'canias_work_center_code', 'Canias’taki benzersiz iş merkezi kodu'),
        ('Departman Kodu (Canias)', 'canias_department_code', 'Canias departman kodu'),
        ('İş Merkezi / Departman Adı', 'name', 'İş merkezi veya departman adı'),
        ('Aktif mi?', 'is_active', '1 = aktif, 0 = pasif'),
    ]),
    ('2_Personel', 'PERSONEL / KULLANICI VERİ GİRİŞİ', [
        ('Personel Sicil No (Canias)', 'personnel_no', 'Canias personel sicil numarası'),
        ('Ad Soyad', 'full_name', 'Personelin ad ve soyadı'),
        ('E-posta', 'email', 'Kurumsal e-posta adresi'),
        ('Rol', 'role', 'admin, operator, maintenance veya sorumlu'),
        ('İş Merkezi Kodu', 'department_id', 'Personelin bağlı olduğu Canias iş merkezi kodu'),
        ('Şifre Hash (Sistem)', 'password_hash', 'Elle yazılmaz; uygulama bcrypt hash üretir'),
        ('Aktif mi?', 'is_active', '1 = aktif, 0 = pasif'),
    ]),
    ('3_Makineler', 'MAKİNE VERİ GİRİŞİ', [
        ('Makine Kodu (Canias)', 'code', 'Canias’taki benzersiz makine kodu; örn. KESIM-01'),
        ('Makine Adı', 'name', 'Makinenin adı veya modeli'),
        ('İş Merkezi Kodu', 'department_id', 'Makinenin bağlı olduğu iş merkezi kodu'),
        ('QR Kodu / Değeri', 'qr_code_value', 'QR içinde bulunacak kod veya URL'),
        ('Canias Varlık / Demirbaş No', 'canias_asset_no', 'Varsa Canias varlık numarası'),
        ('Aktif mi?', 'is_active', '1 = aktif, 0 = pasif'),
    ]),
    ('4_Ariza_Bakim', 'ARIZA / BAKIM VERİ GİRİŞİ', [
        ('Makine Kodu', 'machine_code', 'Arızanın bağlı olduğu Canias makine kodu'),
        ('Bildiren Personel Sicil No', 'reporter_personnel_no', 'Arızayı bildiren personelin Canias sicil no’su'),
        ('Arıza / Bakım Türü', 'malfunction_type', 'Elektrik, mekanik, bakım vb.'),
        ('Öncelik', 'priority', 'low, normal, high veya critical'),
        ('Arıza Açıklaması', 'description', 'Arızanın veya bakım ihtiyacının açıklaması'),
        ('Durum', 'status', 'Açık, İşlemde veya Çözüldü'),
        ('Bildirim Tarihi', 'created_at', 'Tarih-saat; mümkünse UTC'),
        ('Fotoğraf URL', 'photo_url', 'Varsa yüklenen fotoğraf bağlantısı'),
    ]),
]

wb = Workbook()
wb.remove(wb.active)

readme = wb.create_sheet('BAŞLANGIÇ')
readme.sheet_view.showGridLines = False
readme['A1'] = 'SESA-DASH CANIAS VERİ GİRİŞ DOSYASI'
readme['A1'].font = Font(bold=True, color='FFFFFF', size=15)
readme['A1'].fill = HEADER_FILL
readme.merge_cells('A1:D1')
readme['A3'] = 'Verileri hangi sayfalara yazmalısınız?'
readme['A3'].font = Font(bold=True, color='FFFFFF')
readme['A3'].fill = SUBHEADER_FILL
readme.merge_cells('A3:D3')
readme.append([])
readme.append(['Sayfa', 'Ne yazılacak?', 'Önemli alan', 'Doldurma sırası'])
for cell in readme[5]:
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = SUBHEADER_FILL
rows = [
    ['1_Departmanlar', 'Canias iş merkezi ve departman listesi', 'İş Merkezi Kodu', '1'],
    ['2_Personel', 'Personel ve sistem kullanıcıları', 'Personel Sicil No', '2'],
    ['3_Makineler', 'Canias makine listesi ve QR bilgileri', 'Makine Kodu', '3'],
    ['4_Ariza_Bakim', 'Arıza ve bakım kayıtları', 'Makine Kodu + Personel Sicil No', '4'],
]
for row in rows:
    readme.append(row)
readme.append([])
readme.append(['Kullanım notu:', 'Mavi başlıkların altındaki boş satırlara gerçek şirket verilerini yazın.', '', ''])
readme.append(['Kod notu:', 'Kod sütunlarını Excel’de Metin biçiminde tutun; baştaki sıfırlar kaybolmasın.', '', ''])
for col, width in {'A': 24, 'B': 48, 'C': 34, 'D': 16}.items():
    readme.column_dimensions[col].width = width
readme.freeze_panes = 'A6'

for sheet_name, title, columns in sheets:
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False
    headers = [col[0] for col in columns]
    technical = [col[1] for col in columns]
    descriptions = [col[2] for col in columns]
    ws.append([title])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws['A1'].font = Font(bold=True, color='FFFFFF', size=14)
    ws['A1'].fill = HEADER_FILL
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.append(headers)
    ws.append(technical)
    ws.append(descriptions)
    for row_index in (2, 3):
        for cell in ws[row_index]:
            cell.font = Font(bold=True, color='FFFFFF' if row_index == 2 else '555555')
            cell.fill = SUBHEADER_FILL if row_index == 2 else PatternFill('solid', fgColor='E9F0F5')
            cell.alignment = Alignment(wrap_text=True, vertical='center')
            cell.border = Border(bottom=THIN_BLUE)
    for cell in ws[4]:
        cell.font = Font(italic=True, color='555555')
        cell.alignment = Alignment(wrap_text=True, vertical='top')
    ws.row_dimensions[4].height = 38
    ws.freeze_panes = 'A5'
    ws.auto_filter.ref = f'A2:{get_column_letter(len(headers))}104'
    for idx, header in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(idx)].width = max(22, min(38, len(header) + 10))
    for row in ws.iter_rows(min_row=5, max_row=104, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.number_format = '@'
    header_index = {value: index + 1 for index, value in enumerate(headers)}
    for title_text, field, values in [
        ('Rol', 'role', '"admin,operator,maintenance,sorumlu"'),
        ('Öncelik', 'priority', '"low,normal,high,critical"'),
        ('Durum', 'status', '"Açık,İşlemde,Çözüldü"'),
        ('Aktif mi?', 'is_active', '"1,0"'),
    ]:
        if title_text in header_index:
            col = get_column_letter(header_index[title_text])
            validation = DataValidation(type='list', formula1=values, allow_blank=True)
            ws.add_data_validation(validation)
            validation.add(f'{col}5:{col}104')

wb.properties.title = 'SESA-DASH CaniasERP Veri Giriş Şablonu'
wb.properties.creator = 'SESA-DASH'
OUTPUT.parent.mkdir(parents=True, exist_ok=True)
wb.save(OUTPUT)
print(OUTPUT)
